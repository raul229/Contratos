from pathlib import Path
import platform
import warnings
from openpyxl import load_workbook
from docxtpl import DocxTemplate
from subprocess import run, DEVNULL
from utilidades.utils import generar_opciones, mostrar_opciones, completar_fechas

class GestorContratos:
    #quitamos alertas openpyxl
    warnings.simplefilter("ignore", UserWarning)

    def __init__(self, ruta_carpeta_clientes: str | Path, ruta_carpeta_contratos: str | Path) -> None:
        self.ruta_carpeta_clientes = Path(ruta_carpeta_clientes)
        self.ruta_carpeta_contratos = Path(ruta_carpeta_contratos)

        self.contexto: dict | None = None
        self.carpeta_cliente: Path | None = None
        self.editables_cliente: Path | None = None

        self.sistema_operativo: str | None = None
        self.convert = None
        # Ruta de trabajo SIEMPRE como Path
        self.ruta_trabajo: Path | None = None

        self.verificar_os()
        
    def cargar_contexto(self, contexto:dict)->None:
        """
        Carga el diccionario de contexto en el atributo de instancia y completa los datos necesarios.
        Args:
            contexto (dict): El diccionario de contexto a cargar.
        Returns:
            None
        """
        self.contexto=contexto
        self._completar_datos_necesarios()
        

    # --------------------------------------------------
    # LLENAR DATOS PARA CORREO
    # --------------------------------------------------

    def _llenar_valor_celda(self,ws,celda:str,valor:str):
        """
        Llenar un valor en una celda específica
        """
        ws[celda].value = valor
    
    def llenar_datos_correo(self,ruta_excel:Path):
        wb=load_workbook(ruta_excel)
        ws=wb.active
        direccion_instalacion = self.contexto['DOMICILIO_INSTALACION'] if self.contexto['DOMICILIO_INSTALACION'] else f'{self.contexto['DOMICILIO_FISCAL']} - {self.contexto['DISTRITO']}'
        tipo_producto='INTERNET EMPRESAS' if self.contexto['ES_EMPRESAS'] else 'ON NEGOCIOS'
        self._llenar_valor_celda(ws, 'C3', self.contexto['RAZON_SOCIAL'])
        self._llenar_valor_celda(ws, 'C4', self.contexto['RUC'])
        self._llenar_valor_celda(ws, 'C5', self.contexto['RRLL'])
        self._llenar_valor_celda(ws, 'C6', self.contexto['DNI'])
        self._llenar_valor_celda(ws, 'C7', self.contexto['CELULAR_RRLL'])
        self._llenar_valor_celda(ws, 'C8', self.contexto['CORREO_RRLL'])
        self._llenar_valor_celda(ws, 'C9', self.contexto['NOMBRE_OPERATIVO'])
        self._llenar_valor_celda(ws, 'C10', self.contexto['CELULAR_OPERATIVO'])
        self._llenar_valor_celda(ws, 'C11', self.contexto['CORREO_OPERATIVO'])
        self._llenar_valor_celda(ws, 'C12', tipo_producto)
        self._llenar_valor_celda(ws, 'C13', 'X00 MBPS')
        self._llenar_valor_celda(ws, 'C14', 'EECC')
        self._llenar_valor_celda(ws, 'C15', direccion_instalacion)
        self._llenar_valor_celda(ws, 'C16', '')
        self._llenar_valor_celda(ws, 'C17', '')
        self._llenar_valor_celda(ws, 'C18', '')
        
        wb.save(self.carpeta_cliente / ruta_excel.name)
        
    

    def _completar_datos_necesarios(self)->None:

        contexto = self.contexto
        
        lista_respuestas= ['si','s']
        respuesta = input('\n¿El contrato será con firma digital? (s/n):\n> ')
        es_digital = respuesta.lower() in lista_respuestas
        
        respuesta = input('\n¿La venta es un producto empresas? (s/n):\n> ')
        es_empresas = respuesta.lower() in lista_respuestas
        
        contexto = completar_fechas(contexto)

        contexto.update({
            'ES_DIGITAL' : es_digital,
            'ES_EMPRESAS': es_empresas
        })

        #completamos direciones
        direccion = contexto.get('DOMICILIO_FISCAL', '')
        if ' - ' in direccion:
            partes = direccion.split(' - ')
            contexto['DISTRITO'] = partes[-1]
            contexto['DOMICILIO_FISCAL'] = ' - '.join(partes[:-1])

        #representante legal
        nombre = contexto['RRLL']
        dni= contexto['DNI']
        correo = contexto['CORREO_RRLL']
        celular = contexto['CELULAR_RRLL']

        #autocomletamos datos administrativos
        if not (
                contexto['NOMBRE_ADMINISTRATIVO'] and
                contexto['DNI_ADMINISTRATIVO'] and
                contexto['CORREO_ADMINISTRATIVO'] and
                contexto['CELULAR_ADMINISTRATIVO']):
            contexto['NOMBRE_ADMINISTRATIVO'] = nombre
            contexto['DNI_ADMINISTRATIVO'] = dni
            contexto['CORREO_ADMINISTRATIVO'] = correo
            contexto['CELULAR_ADMINISTRATIVO'] = celular

        #autocompletamos datos opetativos
        if not (
                contexto['NOMBRE_OPERATIVO'] and
                contexto['DNI_OPERATIVO'] and
                contexto['CORREO_OPERATIVO'] and
                contexto['CELULAR_OPERATIVO']):
            contexto['NOMBRE_OPERATIVO'] = nombre
            contexto['DNI_OPERATIVO'] = dni
            contexto['CORREO_OPERATIVO'] = correo
            contexto['CELULAR_OPERATIVO'] = celular
            
    # --------------------------------------------------
    # CONSTRUCCIÓN DE RUTA
    # --------------------------------------------------

    def construir_ruta_trabajo(self) -> None:
        """
        Construye la ruta de trabajo seleccionando las carpetas interactivamente
        """
        ruta_actual = self.ruta_carpeta_contratos

        while True:
            opciones = generar_opciones(ruta_actual)

            if '0. Obligatorios' in opciones:
                break

            opcion = mostrar_opciones(opciones)
            ruta_actual = ruta_actual / opcion

        self.ruta_trabajo = ruta_actual

    # --------------------------------------------------
    # SISTEMA OPERATIVO
    # --------------------------------------------------

    def verificar_os(self) -> None:
        if platform.system() == 'Windows':
            from docx2pdf import convert
            self.convert = convert
            self.sistema_operativo = 'Windows'
        else:
            self.sistema_operativo = 'Linux'

    # --------------------------------------------------
    # CARPETAS CLIENTE
    # --------------------------------------------------

    def crear_carpeta_cliente(self) -> None:
        self.carpeta_cliente = self.ruta_carpeta_clientes / f'{self.contexto["RAZON_SOCIAL"]}-{self.contexto["RUC"]}'
        self.editables_cliente = self.carpeta_cliente / 'EDITABLES'

        self.carpeta_cliente.mkdir(exist_ok=True)
        self.editables_cliente.mkdir(exist_ok=True)
        (self.carpeta_cliente / 'DOC ADICIONALES').mkdir(exist_ok=True)

    # --------------------------------------------------
    # LLENADO DE PLANTILLAS
    # --------------------------------------------------

    def _llenar_docx(self, ruta_plantilla: Path) -> None:
        docx = DocxTemplate(ruta_plantilla)
        docx.render(self.contexto)
        docx.save(self.editables_cliente / ruta_plantilla.name)

    def _llenar_xlsx(self, ruta_plantilla: Path) -> None:

        #llenamos campos validados
        wb = load_workbook(ruta_plantilla)
        ws = wb.active
        ws['F44'].value = self.contexto['RUC']
        ws['F45'].value = self.contexto['RAZON_SOCIAL']
        ws['F55'].value = self.contexto['DOMICILIO_INSTALACION'] if self.contexto['DOMICILIO_INSTALACION'] else f'{self.contexto['DOMICILIO_FISCAL']} - {self.contexto['DISTRITO']}'
        ws['C69'].value = self.contexto['DIA']
        ws['E69'].value = self.contexto['NOMBRE_MES']
        ws['G69'].value = self.contexto['ANIO']
        ws['K76'].value = self.contexto['RRLL']
        ws['K77'].value = 'GERENTE GENERAL'
        ws['K78'].value = self.contexto['DNI']

        wb.save(self.editables_cliente / ruta_plantilla.name)


    def llenar_plantilla(self, nombre_plantilla: str | list[str]) -> None:
        
        self.crear_carpeta_cliente()

        if isinstance(nombre_plantilla, list):
            rutas = [
                self.ruta_trabajo / nombre
                for nombre in nombre_plantilla
            ]
        else:
            rutas = [self.ruta_trabajo / nombre_plantilla]

        for ruta in rutas:
            if ruta.suffix == '.docx':
                self._llenar_docx(ruta)
            elif ruta.suffix in ('.xlsx', '.xlsm'):
                self._llenar_xlsx(ruta)
            else:
                raise ValueError(f'Formato no soportado: {ruta.name}')

    # --------------------------------------------------
    # CONVERSIÓN A PDF
    # --------------------------------------------------

    def convertir_a_pdf(self, nombre_archivo: str) -> None:

        #se convierten a pdf desde los editables del cliente
        ruta_origen = self.editables_cliente / nombre_archivo

        if self.sistema_operativo == 'Linux':
            run([
                'libreoffice',
                '--headless',
                '--convert-to', 'pdf',
                '--outdir', str(self.carpeta_cliente),
                str(ruta_origen)
            ], check=True, stdout=DEVNULL, stderr=DEVNULL)

        else:
            if ruta_origen.suffix != '.docx':
                raise ValueError('En Windows solo se convierten archivos .docx')

            ruta_pdf = self.carpeta_cliente / ruta_origen.with_suffix('.pdf').name
            self.convert(str(ruta_origen), str(ruta_pdf))
            
        
