from abc import ABC, abstractmethod
from pathlib import Path
import platform
import warnings

from openpyxl import load_workbook
from docxtpl import DocxTemplate

from utilidades.utils import (
    convertir_a_pdf,
    generar_opciones,
    mostrar_opciones,
    completar_fechas,
)


class GeneradorContratos(ABC):
    """
    Clase base abstracta para todos los generadores de contratos.
    Define el flujo común:
        1. Cargar contexto
        2. Construir ruta de plantillas
        3. Llenar plantillas (docx / xlsx / pdf según la subclase)
        4. Convertir a PDF
        5. Llenar excel de datos de correo
    """

    warnings.simplefilter("ignore", UserWarning)

    def __init__(
        self,
        ruta_carpeta_clientes: str | Path,
        ruta_carpeta_contratos: str | Path,
    ) -> None:
        self.ruta_carpeta_clientes = Path(ruta_carpeta_clientes)
        self.ruta_carpeta_contratos = Path(ruta_carpeta_contratos)

        self.contexto: dict | None = None
        self.carpeta_cliente: Path | None = None
        self.editables_cliente: Path | None = None

        self.sistema_operativo: str | None = None
        self.convert = None
        self.ruta_trabajo: Path | None = None

        self.verificar_os()

    # --------------------------------------------------
    # API PÚBLICA (plantilla del flujo)
    # --------------------------------------------------

    def ejecutar(self, contexto: dict) -> None:
        """Ejecuta el flujo completo de generación de contratos."""
        self.cargar_contexto(contexto)
        self.construir_ruta_trabajo()
        self.preparar_carpetas_cliente()

        plantillas = self.obtener_plantillas()
        self.llenar_plantillas(plantillas)

        self.convertir_a_pdf(plantillas)

        if self.usa_excel_correo():
            self.llenar_datos_correo(self.ruta_excel_correo())

        self.post_proceso()

    # --------------------------------------------------
    # HOOKS QUE LAS SUBCLASES PUEDEN SOBREESCRIBIR
    # --------------------------------------------------

    @abstractmethod
    def nombre_operador(self) -> str:
        """Nombre legible del operador, usado para carpetas/logs."""

    def usa_excel_correo(self) -> bool:
        """Indica si este operador requiere llenar el excel de datos de correo."""
        return True

    def ruta_excel_correo(self) -> Path | None:
        """Ruta al excel de datos de correo (si aplica)."""
        return None

    def post_proceso(self) -> None:
        """Hook final opcional (logs, notificaciones, etc.)."""

    # --------------------------------------------------
    # CONTEXTO
    # --------------------------------------------------

    def cargar_contexto(self, contexto: dict) -> None:
        self.contexto = contexto
        self._completar_datos_necesarios()

    def _completar_datos_necesarios(self) -> None:
        contexto = completar_fechas(self.contexto or {})
        self._completar_datos_particulares(contexto)
        self.contexto = contexto

    def _completar_datos_particulares(self, contexto: dict) -> None:
        """Hook que las subclases usan para añadir datos específicos."""

    # --------------------------------------------------
    # RUTA DE TRABAJO
    # --------------------------------------------------

    def construir_ruta_trabajo(self) -> None:
        ruta_actual = self.ruta_carpeta_contratos

        while True:
            opciones = generar_opciones(ruta_actual)
            if self._es_ruta_valida(opciones):
                break
            opcion = mostrar_opciones(opciones)
            ruta_actual = ruta_actual / opcion

        self.ruta_trabajo = ruta_actual

    def _es_ruta_valida(self, opciones: list[str]) -> bool:
        """Las subclases definen cómo se reconoce una ruta con plantillas."""
        return '0. Obligatorios' in opciones

    # --------------------------------------------------
    # SISTEMA OPERATIVO
    # --------------------------------------------------

    def verificar_os(self) -> None:
        if platform.system() == 'Windows':
            from docx2pdf import convert
            self.convert = convert
            self.sistema_operativo = 'Windows'
        else:
            self.sistema_operativo = 'Linux'

    # --------------------------------------------------
    # CARPETAS DEL CLIENTE
    # --------------------------------------------------

    def preparar_carpetas_cliente(self) -> None:
        self.crear_carpeta_cliente()

    def crear_carpeta_cliente(self) -> None:
        self.carpeta_cliente = (
            self.ruta_carpeta_clientes
            / f'{self.contexto["RAZON_SOCIAL"]}-{self.contexto["RUC"]}'
        )
        self.editables_cliente = self.carpeta_cliente / 'EDITABLES'

        self.carpeta_cliente.mkdir(exist_ok=True)
        self.editables_cliente.mkdir(exist_ok=True)
        (self.carpeta_cliente / 'DOC ADICIONALES').mkdir(exist_ok=True)

    # --------------------------------------------------
    # PLANTILLAS
    # --------------------------------------------------

    def obtener_plantillas(self) -> list[Path]:
        """Devuelve la lista de rutas de plantillas a llenar."""
        raise NotImplementedError

    def llenar_plantillas(self, plantillas: list[Path]) -> None:
        for ruta in plantillas:
            self.llenar_una_plantilla(ruta)

    def llenar_una_plantilla(self, ruta: Path) -> None:
        """Lógica común: enruta según extensión."""
        if ruta.suffix == '.docx':
            self._llenar_docx(ruta)
        elif ruta.suffix in ('.xlsx', '.xlsm'):
            self._llenar_xlsx(ruta)
        elif ruta.suffix == '.pdf':
            self._llenar_pdf(ruta)
        else:
            raise ValueError(f'Formato no soportado: {ruta.name}')

    # Las subclases pueden sobreescribir estos métodos si necesitan
    # un comportamiento particular.
    def _llenar_docx(self, ruta_plantilla: Path) -> None:
        docx = DocxTemplate(ruta_plantilla)
        docx.render(self.contexto)
        docx.save(self.editables_cliente / ruta_plantilla.name)

    def _llenar_xlsx(self, ruta_plantilla: Path) -> None:
        """
        Motor genérico: aplica un mapa campo -> (hoja, celda) sobre
        cualquier plantilla xlsx/xlsm. El mapa se obtiene de
        `coordenadas_para_xlsx(ruta)`. Si no hay mapa, se conserva el
        archivo sin cambios en EDITABLES.
        """
        import shutil

        coords = self.coordenadas_para_xlsx(ruta_plantilla)
        if not coords:
            shutil.copy(
                ruta_plantilla,
                self.editables_cliente / ruta_plantilla.name,
            )
            return

        destino = self.editables_cliente / ruta_plantilla.name
        shutil.copy(ruta_plantilla, destino)

        wb = load_workbook(destino, keep_vba=True)
        for campo, (hoja, celda) in coords.items():
            valor = self._valor_para_campo(campo)
            if valor in (None, ''):
                continue
            wb[hoja][celda] = valor

        if hasattr(self, '_post_procesar_xlsx'):
            self._post_procesar_xlsx(wb)

        wb.save(destino)

    def coordenadas_para_xlsx(self, ruta_plantilla: Path) -> dict:
        """
        Devuelve {campo: (hoja, celda)} para la plantilla indicada.
        Por defecto vacío; las subclases lo sobreescriben.
        """
        return {}

    def _valor_para_campo(self, campo: str):
        """
        Resolución del valor a escribir en una celda. Prioridad:
            1. texto fijo (string que empieza con '=' no aplica)
            2. contexto del cliente
        Las subclases pueden sobreescribir para añadir datos propios.
        """
        if campo in (self.contexto or {}):
            return self.contexto[campo]
        return None

    def _llenar_pdf(self, ruta_plantilla: Path) -> None:
        raise NotImplementedError(
            f'{self.__class__.__name__} debe implementar _llenar_pdf '
            f'para plantillas PDF.'
        )

    def _direccion_instalacion(self) -> str:
        di = self.contexto.get('DOMICILIO_INSTALACION')
        if di:
            return di
        return (
            f"{self.contexto.get('DOMICILIO_FISCAL', '')}"
            f" - {self.contexto.get('DISTRITO', '')}"
        ).strip(' -')

    # --------------------------------------------------
    # CONVERSIÓN A PDF
    # --------------------------------------------------

    def convertir_a_pdf(self, plantillas: list[Path]) -> None:
        for ruta in plantillas:
            self._convertir_archivo_a_pdf(ruta)

    def _convertir_archivo_a_pdf(self, ruta: Path) -> None:
        if ruta.suffix == '.pdf':
            return  # ya es PDF, no se convierte

        ruta_origen = self.editables_cliente / ruta.name
        pdf = convertir_a_pdf(ruta_origen, self.sistema_operativo)
        if pdf is not None and pdf.parent != self.carpeta_cliente:
            destino = self.carpeta_cliente / pdf.name
            pdf.replace(destino)

    # --------------------------------------------------
    # EXCEL DE CORREO
    # --------------------------------------------------

    def _llenar_valor_celda(self, ws, celda: str, valor) -> None:
        ws[celda].value = valor

    def llenar_datos_correo(self, ruta_excel: Path) -> None:
        wb = load_workbook(ruta_excel)
        ws = wb.active

        direccion_instalacion = self._direccion_instalacion()
        tipo_producto = self._tipo_producto_correo()

        self._llenar_valor_celda(ws, 'C3', self.contexto['RAZON_SOCIAL'])
        self._llenar_valor_celda(ws, 'C4', self.contexto['RUC'])
        self._llenar_valor_celda(ws, 'C5', self.contexto['RRLL'])
        self._llenar_valor_celda(ws, 'C6', self.contexto['DNI'])
        self._llenar_valor_celda(ws, 'C7', self.contexto['CELULAR_RRLL'])
        self._llenar_valor_celda(ws, 'C8', self.contexto['CORREO_RRLL'])
        self._llenar_valor_celda(ws, 'C9', self.contexto['NOMBRE_OPERATIVO'])
        self._llenar_valor_celda(ws, 'C10', self.contexto['CELULAR_OPERATIVO'])
        self._llenar_valor_celda(ws, 'C11', self.contexto['CORREO_OPERATIVO'])
        self._llenar_valor_celda(ws, 'C12', tipo_producto)
        self._llenar_valor_celda(ws, 'C13', self._velocidad_etiqueta_correo())
        self._llenar_valor_celda(ws, 'C14', 'EECC')
        self._llenar_valor_celda(ws, 'C15', direccion_instalacion)
        self._llenar_valor_celda(ws, 'C16', '')
        self._llenar_valor_celda(ws, 'C17', '')
        self._llenar_valor_celda(ws, 'C18', '')

        wb.save(self.carpeta_cliente / ruta_excel.name)

    def _tipo_producto_correo(self) -> str:
        return (
            'INTERNET EMPRESAS'
            if self.contexto.get('ES_EMPRESAS')
            else 'ON NEGOCIOS'
        )

    def _velocidad_etiqueta_correo(self) -> str:
        return 'X00 MBPS'
